import vpython as vp 
import openpyxl as op

print("hello")

try:
    workbook = op.Workbook()  # Temp creates workbook in memroy but its not saved to harddrive
    sheet = workbook.active  # Sets this workbook as the one to write too, gets the first sheet of workbook

    print("sheet name: " + sheet.title)

    sheet.title = "Sheet 1"

    print("sheet rename: " + sheet.title)

    # Note: The first row or column integer is 1, not 0. Cell object is created by 
    # using sheet object's cell() method. 
    c1 = sheet.cell(row = 1, column = 1) 
    # writing values to cells 
    c1.value = "ANKIT"

    c2 = sheet.cell(row= 1 , column = 2) 
    c2.value = "RAI"

    # Once have a Worksheet object, one can access a cell object by its name also. 
    # A2 means column = 1 & row = 2. 
    c3 = sheet['A2'] 
    c3.value = "RAHUL"

    # B2 means column = 2 & row = 2. 
    c4 = sheet['B2'] 
    c4.value = "RAI"

    workbook.save("C:\\Users\\alexi\\Documents\\Programming\\VC Code\\Python\\Truss Simulation Project\\testingsile.xlsx")
    
    workbook.close
    
    box = vp.sphere(pos=vp.vec(0,-1,0))

except:
    print("Some general problem occured")

